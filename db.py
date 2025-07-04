import openpyxl
import datetime
import uuid
import json
import pytz
from PIL import Image
import requests

class Database:
    def __init__(self, path):
        self.path = path
        self.workbook = openpyxl.load_workbook(path)
        self.sessions = {}

        self.load_db()

    def load_db(self):
        self.users = self.load_data(self.workbook, 'users', User)
        self.tokens = self.load_data(self.workbook, 'tokens', Token)
        self.transactions = self.load_data(self.workbook, 'transactions', Transaction)
        self.listings = self.load_data(self.workbook, 'listings', Listing)
        self.fish_catches = self.load_data(self.workbook, 'fish_catches', FishCatches)
        self.submissions = self.load_data(self.workbook, 'submissions', Submission)

    def create_account(self, username, password, nickname):
        worksheet=self.workbook['users']       
        worksheet.append([
            max(self.users.keys() or [0])+1,
            username,
            password,
            nickname,
            datetime.datetime.now().isoformat(),
            False
        ])
        self.workbook.save(self.path)
        self.load_db()

    def write_transaction(self, user_from, user_to, amount=0, token=None):
        worksheet=self.workbook['transactions']
        worksheet.append([
            max(self.transactions.keys())+1,
            datetime.datetime.now().isoformat(),
            user_from,
            user_to,
            amount,
            token
        ])
        self.workbook.save(self.path)
        self.load_db()

    def write_listing(self, seller_id, token_id, amount=None):
        worksheet=self.workbook['listings']       
        worksheet.append([
            max(self.listings.keys())+1,
            datetime.datetime.now().isoformat(),
            seller_id,
            token_id,
            amount,
        ])
        self.workbook.save(self.path)
        self.load_db()

    def write_fish_catch(self, species, weight_lbs, length_in, angler_id, location_id):
        worksheet=self.workbook['fish_catches']       
        worksheet.append([
            max(self.fish_catches.keys() or [0])+1,
            datetime.datetime.now().isoformat(),
            species,
            weight_lbs,
            length_in,
            angler_id,
            location_id
        ])
        self.workbook.save(self.path)
        self.load_db()

    def write_new_token_submission(self, token_note, token_url, token_hash, token_author_id):

        # First create a new token record
        # Token is disabled=True
        token_id = max(self.tokens.keys() or [0])+1

        worksheet=self.workbook['tokens']
        worksheet.append([
            token_id,
            datetime.datetime.now().isoformat(),
            token_note,
            token_url,
            True,
            token_hash
        ])
        self.workbook.save(self.path)

        # Then create submission for that token
        # reviewed = False
        worksheet=self.workbook['submissions']       
        worksheet.append([
            max(self.submissions.keys() or [0])+1,
            datetime.datetime.now().isoformat(),
            token_author_id,
            token_id,
            False
        ])
        self.workbook.save(self.path)

        self.load_db()

    def submission_approve(self, submission_id):
        # Submission is approved
        # 1. Set Submission reviewed = True
        # 2. Set Token disabled = False
        # 3. Send Token from System to User
        submission = self.submissions[submission_id]

        # 1. Set Submission reviewed = True
        worksheet=self.workbook['submissions']
        worksheet.cell(row=submission_id+1, column=5).value = True
        self.workbook.save(self.path)

        # 2. Set Token disabled = False
        worksheet=self.workbook['tokens']
        worksheet.cell(row=submission.token.id+1, column=5).value = False
        self.workbook.save(self.path)

        # 3. Send Token from System to User
        self.write_transaction(
            user_from=0,
            user_to=submission.author.id,
            token=submission.token.id
        )

        self.workbook.save(self.path)
        self.load_db()

    def submission_deny(self, submission_id):
        # Submission is denied
        # 1. Set Submission reviewed = True
        # 2. Delete Token
        worksheet=self.workbook['submissions']
        worksheet.cell(row=submission_id+1, column=5).value = True
        self.workbook.save(self.path)
        self.load_db()

    def update_user_password(self, user, password):
        # Update user password
        worksheet=self.workbook['users']
        worksheet.cell(row=user.id+2, column=3).value = password
        self.workbook.save(self.path)
        self.load_db()

    def load_data(self, workbook, worksheet, pattern):
        first = True
        headers = None
        data = {}
        for i, row in enumerate(workbook[worksheet].iter_rows(values_only=True)):
            if first:
                headers = row
                first = False
            else:
                rowdict = dict()
                for i, header in enumerate(headers):
                    rowdict[header] = row[i]
                obj = pattern(rowdict, self)
                data[obj.id] = obj
        return data
        
    def get_user(self, username):
        for id,user in self.users.items():
            if user.username == username:
                return user
        return None
    
    def get_token(self, token_id):
        for id,token in self.tokens.items():
            if token.id == token_id:
                return token
        return None

    def get_all_token_titles(self):
        token_titles = []
        for id,token in self.tokens.items():
            token_titles.append(token.note)
        return token_titles
    
    def get_all_token_urls(self):
        token_urls = []
        for id,token in self.tokens.items():
            token_urls.append(token.note)
        return token_urls
    
    def get_all_token_hashes(self):
        token_hashes = []
        for id,token in self.tokens.items():
            token_hashes.append(token.hash)
        return token_hashes

    def start_session(self, username):
        session = Session(username)
        self.sessions[username] = session
        return session

    def end_session(self, username):
        self.sessions.pop(username)

    def check_session(self, token):
        for id,session in self.sessions.items():
            if session.token == token:
                return self.get_user(session.username)
        return None

    def user_list(self, user):
        users = []
        for id,u in self.users.items():
            if u.username == 'system':
                continue
            if u.username == user.username:
                continue
            users.append(u.username)
        return users
    
    def all_usernames(self):
        usernames = []
        for id,user in self.users.items():
            usernames.append(user.username)
        return usernames
    
    def all_nicknames(self):
        nicknames = []
        for id,user in self.users.items():
            nicknames.append(user.nickname)
        return nicknames

    def transaction_list(self):
        transactions = []
        for i,t in self.transactions.items():
            transactions.append(t.to_dict())
        transactions.reverse()
        return transactions

    def for_sale(self):
        for_sale = {}
        for id,listing in self.listings.items():
            if listing.amount:
                for_sale[listing.token.id] = listing
            else:
                for_sale.pop(listing.token.id)
        return for_sale
    
    def pending_submissions(self):
        subs = []
        
        for id, sub in self.submissions.items():
            if not sub.reviewed:
                subs.append(sub)
        
        return subs

class Object:
    def __init__(self, d, db):
        self.db = db
        for key in d:
            setattr(self, key, d[key])

    def __iter__(self):
        for p in dir(self):
            if not p.startswith('__'):
                yield p, getattr(self, p)

    def __repr__(self):
        d = {}
        for p in dir(self):
            if p.startswith('__'):
                continue
            if p == 'db':
                continue

            if type(getattr(self, p)) == User:
                d[p] = dict(getattr(self, p))
            else:
                d[p] = getattr(self, p)

        return json.dumps(d)

class User(Object):

    @property
    def awards(self):

        awards = []

        # Achievement 1 : Most LBC
        a1 = True
        for id,user in self.db.users.items():
            if id == 0:
                continue
            if user.balance > self.balance:
                a1 = False
        if a1:
            awards.append(
                Achievement('Most LBC', 'award1.png')
            )

        # Achievement 2 : Most NFT
        a2 = True
        for id,user in self.db.users.items():
            if id == 0:
                continue
            if len(user.tokens) > len(self.tokens):
                a2 = False
        if a2:
            awards.append(
                Achievement('Most NFT', 'award2.png')
            )

        # Achievement 3 : 100% Collection Log
        a3 = True
        from activities.fishing import Fishing
        fishing = Fishing(self.db)
        if len(self.fish_species) < len(fishing.species):
            a3 = False 
        if a3:
            awards.append(
                Achievement('All Species', 'award3.png')
            )

        awards.append(Achievement('', 'award0.png'))

        return awards

    @property
    def tokens(self):
        tokens = {}
        for id, transaction in self.db.transactions.items():

            if getattr(transaction, 'to') is self.id and transaction.token:
                tokens[transaction.token] = self.db.tokens[transaction.token].to_dict()

            if getattr(transaction, 'from') is self.id and transaction.token:
                if transaction.token in tokens:
                    tokens.pop(transaction.token)
        
        tokens = dict(reversed(tokens.items()))
        return tokens

    @property
    def balance(self):
        balance = 0

        for id, transaction in self.db.transactions.items():

            if getattr(transaction, 'to') is self.id and transaction.amount:
                balance += transaction.amount

            if getattr(transaction, 'from') is self.id and transaction.amount:               
                if getattr(transaction, 'to') == getattr(transaction, 'from') and self.id == 0:
                    continue

                balance -= transaction.amount
        
        return balance
    
    @property
    def transactions(self):
        transactions = []

        for id, transaction in self.db.transactions.items():

            if getattr(transaction, 'to') is self.id:
                transactions.append(transaction.to_dict())
                continue

            if getattr(transaction, 'from') is self.id:
                transactions.append(transaction.to_dict())
                continue

        transactions.reverse()

        return transactions

    @property
    def submissions(self):
        submissions = []
        for id, submission in self.db.submissions.items():
            if submission.author_id == self.id:
                submissions.append(submission)
        submissions.reverse()
        return submissions

    @property
    def fish_catches(self):
        fish_catches = []
        for id, fish in self.db.fish_catches.items():
            if fish.angler == self.id:
                fish_catches.append(fish)

        fish_catches.reverse()
        return fish_catches

    def fish_catches_species(self, species):
        # List of catches of a specific species
        fish_catches = []

        for fish in self.fish_catches:
            if fish.species == species:
                fish_catches.append(fish)

        return fish_catches
    
    def fish_catches_species_stats(self, species):
        # Generate statistics for fish caught of a particular species
        length_in_longest = 0
        length_in_shortest = 0
        weight_lbs_heaviest = 0
        weight_lbs_lightest = 0
        number_caught = 0

        for fish in self.fish_catches_species(species):

            number_caught += 1

            if not length_in_longest or fish.length_in > length_in_longest:
                length_in_longest = fish.length_in

            if not length_in_shortest or fish.length_in < length_in_shortest:
                length_in_shortest = fish.length_in

            if not weight_lbs_heaviest or fish.weight_lbs > weight_lbs_heaviest:
                weight_lbs_heaviest = fish.weight_lbs

            if not weight_lbs_lightest or fish.weight_lbs < weight_lbs_lightest:
                weight_lbs_lightest = fish.weight_lbs

        return {
            'length_in_longest': length_in_longest,
            'length_in_shortest': length_in_shortest,
            'weight_lbs_heaviest': weight_lbs_heaviest,
            'weight_lbs_lightest': weight_lbs_lightest,
            'number_caught': number_caught
        }

    @property
    def fish_species(self):
        # List of unique fish species user has caught
        species = []
        for fish in self.fish_catches:
            if not fish.species in species:
                species.append(fish.species)
        return species

    @property
    def fished_today(self):
        # Indicate the count of fish user has caught today
        eastern = pytz.timezone('US/Eastern')
        fish_caught_today = 0

        # No catch history means user has never fished before
        if len(self.fish_catches) == 0:
            return fish_caught_today

        for fish in self.fish_catches:
            fish_caught_ts = datetime.datetime.fromisoformat(fish.timestamp).replace(tzinfo=datetime.UTC).astimezone(eastern)

            # "Yesterday" is ditermined by the calendar day.
            if not fish_caught_ts.date() < datetime.datetime.today().astimezone(eastern).date():
                fish_caught_today += 1
        
        return fish_caught_today

    def fish_species_from_location(self, location):
        # Return list of unique species caught at a location
        species = []
        for fish in self.fish_catches:
            if fish.location_id == location.id and fish.species not in species:
                species.append(fish.species)
        return species

    def fish_species_complete_from_location(self, location):
        # The % of possible speices user has caught from a location
        percent_complete = len(self.fish_species_from_location(location)) / len(location.species)
        return int(percent_complete*100)

    @property
    def submissions_this_week(self):
        # Indicate the count of fish user has caught today
        eastern = pytz.timezone('US/Eastern')
        submissions_this_week = 0

        # No submissions means never submitted before
        if len(self.submissions) == 0:
            return submissions_this_week

        for submission in self.submissions:
            submission_ts = datetime.datetime.fromisoformat(submission.created_at).replace(tzinfo=datetime.UTC).astimezone(eastern)

            # "Yesterday" is ditermined by the calendar day.
            if submission_ts.date() > datetime.datetime.today().astimezone(eastern).date() - datetime.timedelta(weeks=1):
                submissions_this_week += 1
        
        return submissions_this_week

    def to_dict(self):
        d = {
            'id': self.id,
            'username': self.username,
            'nickname': self.nickname,
            'created_at': self.created_at,
            'balance': self.balance,
            'transactions': self.transactions,
            'tokens': self.tokens
        }
        return d

class Token(Object):

    next_hash_check = None

    @property
    def url(self):    
        return self.link
        # This method is supped to check the linked image hasn't changes since it was approved.
        # If the image changes to something else other than the approved version it's blanked out.
        # However, in-practice checking every single token make the app run slow.
        #
        # if self.next_hash_check and datetime.datetime.now() < self.next_hash_check:
        #     # If token hash was just checked and it's fine return the link
        #     return self.link
        # else:
        #     # If token hasn't been checked. Check it.
        #     if hash_img(self.link) == self.hash:
        #         # If the hash matches set a future check ts and return the link
        #         self.next_hash_check = datetime.datetime.now() + datetime.timedelta(days=1)
        #         return self.link
        #     else:
        #         # If the hash doesn't match disable it and return blank
        #         self.disabled = True
        #         return '/static/blank.jpg'

    @property
    def owner(self):
        owner = self.db.users[0]
        for id, transaction in self.db.transactions.items():

            if transaction.amount:
                continue

            if transaction.token is self.id:
                owner = self.db.users[transaction.to]

        return owner
    
    @property
    def transactions(self):
        t = []
        for id, transaction in self.db.transactions.items():
            if transaction.token is self.id:
                t.append(transaction)
        t.reverse()
        return t

    @property
    def for_sale(self):
        if self.id in self.db.for_sale():
            return True
        else:
            return False

    @property
    def listing(self):
        if self.for_sale:
            return self.db.for_sale()[self.id]
        else:
            return None

    @property
    def submission(self):
        for id, sub in self.db.submissions.items():
            if sub.token_id == self.id:
                return sub

    def to_dict(self):
        d = {
            'id': self.id,
            'created_at': self.created_at,
            'note': self.note,
            'url': self.url,
            'for_sale': self.for_sale
        }
        return d

class Transaction(Object):

    @property
    def user_from(self):
        return self.db.users[getattr(self, 'from')]

    @property
    def user_to(self):
        return self.db.users[getattr(self, 'to')]

    def to_dict(self):
        d = {
            'id': self.id,
            'timestamp': self.timestamp,
            'amount': self.amount,
            'token': self.token,
            'user_from_id': self.user_from.id,
            'user_from_username': self.user_from.username,
            'user_from_nickname': self.user_from.nickname,
            'user_to_id': self.user_to.id,
            'user_to_username': self.user_to.username,
            'user_to_nickname': self.user_to.nickname,
            'token_url': None,
            'token_note': None

        }
        if self.token:
            d['token_url'] = self.db.tokens[self.token].url
            d['token_note'] = self.db.tokens[self.token].note

        return d

class Achievement:
    def __init__(self, name, icon):
        self.name = name
        self.icon = icon

class Listing(Object):

    @property
    def seller(self):
        return self.db.users[getattr(self, 'seller_id')]
    
    @property
    def token(self):
        return self.db.tokens[getattr(self, 'token_id')]

class Submission(Object):

    @property
    def author(self):
        return self.db.users[self.author_id]
    
    @property
    def token(self):
        return self.db.tokens[self.token_id]

class Session:
    def __init__(self, username):
        self.username = username
        self.expires = datetime.datetime.now() + datetime.timedelta(days=30)
        self.token = str(uuid.uuid4())

class FishCatches(Object):
    pass

def hash_img(url):
    # Take a PIL image and compute a "hash" representing the state of the image.
    # https://stackoverflow.com/questions/49689550/simple-hash-of-pil-image

    rsp = requests.get(
        url, 
        stream=True,
        headers={
            'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/136.0.0.0 Safari/537.36'
        }
    )

    img = Image.open(rsp.raw)
    img = img.resize((20, 20))
    img = img.convert("P")
    pixel_data = list(img.getdata())
    avg_pixel = sum(pixel_data)/len(pixel_data)
    bits = "".join(['1' if (px >= avg_pixel) else '0' for px in pixel_data])
    hex_representation = str(hex(int(bits, 2)))[2:][::-1].upper()
    return hex_representation
